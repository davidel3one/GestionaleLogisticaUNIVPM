import csv
from concurrent.futures import Future
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session, aliased, sessionmaker

from gestionale_logistica.concorrenza import esegui_in_background
from gestionale_logistica.database.base import SessionLocal
from gestionale_logistica.database.crud_base import CRUDBase
from gestionale_logistica.database.enums import CategoriaConsegna, StatoOrdine, StatoViaggio
from gestionale_logistica.database.models import (
    Camion,
    ComposizioneSquadra,
    Dipendente,
    EsitoConsegna,
    Ordine,
    Viaggio,
    report_ordini,
)
from gestionale_logistica.logistica.geocoding import geocodifica_comune

COLONNE_ATTESE = ["ID_Ordine", "Cliente", "Indirizzo", "Categoria", "Peso", "Volume", "Provincia"]

ordine = CRUDBase[Ordine](Ordine)
viaggio = CRUDBase[Viaggio](Viaggio)

FILTRO_TUTTI = "Tutti"

# Motivi di rifiuto RF11 legati all'idoneita' categoria<->risorsa (non a peso/volume residuo):
# estratti come costanti (invece che inline in valida_ordine_per_viaggio) cosi' la GUI puo'
# distinguerli senza duplicare il testo — vedi ManualeTab._aggiungi_ordine, che li mostra come
# toast invece dell'alert sotto la tabella "Aggiungi ordine".
MOTIVO_SPONDA_IDRAULICA_MANCANTE = "Il camion non ha la sponda idraulica necessaria per ordini di categoria Big"
MOTIVO_CERTIFICAZIONE_GAS_MANCANTE = "Nessun membro della squadra ha la certificazione gas necessaria"
# RF12: motivo di rifiuto in avvia_composizione_viaggio quando la composizione e' gia' occupata
# da un altro viaggio nella stessa data — estratto come costante cosi' la GUI puo' distinguerlo
# senza duplicare il testo, vedi AssistitaTab._avvia_composizione che lo mostra come toast.
MOTIVO_COMPOSIZIONE_OCCUPATA = "Composizione gia' occupata in questa data"


def _normalizza_filtro_multiplo(valore: str | list[str] | None, sentinella: str | None) -> set[str] | None:
    """Vedi _normalizza_filtro_multiplo in gestore_dipendenti.py: stessa logica, duplicata perche'
    i gestori non condividono un modulo utils comune."""
    if valore is None or valore == sentinella:
        return None
    if isinstance(valore, str):
        return {valore}
    return set(valore) or None

# Etichette italiane per la lista Ordini: gli enum StatoOrdine hanno valori CamelCase pensati per
# la persistenza, non per la UI (es. RICEVUTO -> "Da pianificare", non "Ricevuto" - coerente col
# mockup, dove un ordine appena importato/non ancora agganciato a un viaggio si legge come "in
# attesa di essere pianificato", non come "ricevuto e basta").
STATO_ORDINE_LABELS: dict[StatoOrdine, str] = {
    StatoOrdine.RICEVUTO: "Da pianificare",
    StatoOrdine.PIANIFICATO: "Pianificato",
    StatoOrdine.IN_CONSEGNA: "In consegna",
    StatoOrdine.COMPLETATO: "Consegnato",
    StatoOrdine.FALLITO: "Fallito",
}

# Etichette italiane per la lista Viaggi: gli enum StatoViaggio hanno valori CamelCase
# ("InComposizione") pensati per la persistenza, non per la UI. Non c'e' un campo "stato" derivato
# gia' in italiano come per Dipendenti/Camion (lo stato del viaggio e' gia' un campo diretto del
# modello), quindi serve una mappa esplicita anche per il testo, non solo per il colore.
STATO_VIAGGIO_LABELS: dict[StatoViaggio, str] = {
    StatoViaggio.IN_COMPOSIZIONE: "In composizione",
    StatoViaggio.PIANIFICATO: "Pianificato",
    StatoViaggio.IN_CORSO: "In corso",
    StatoViaggio.COMPLETATO: "Completato",
    StatoViaggio.ANNULLATO: "Annullato",
}

ORDINA_PER_PARTENZA = "data_partenza_prevista"
ORDINA_PER_ARRIVO = "data_arrivo_prevista"


def _righe_da_xlsx(percorso_file: Path) -> tuple[list[str] | None, list[dict[str, str | None]]]:
    """Legge un file Excel (RF9) e lo riporta alla stessa forma di csv.DictReader (header +
    righe come dict di stringhe), cosi' la validazione in importa_ordini resta unica per
    entrambi i formati. Le celle mancanti in coda a una riga corta diventano None, come il
    restval di DictReader per una riga CSV con meno colonne dell'header."""
    wb = load_workbook(percorso_file, read_only=True, data_only=True)
    try:
        righe_grezze = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()
    if not righe_grezze:
        return None, []

    header = [str(valore) if valore is not None else "" for valore in righe_grezze[0]]
    righe = []
    for row in righe_grezze[1:]:
        valori = list(row) + [None] * (len(header) - len(row))
        righe.append({col: None if v is None else str(v) for col, v in zip(header, valori)})
    return header, righe


@dataclass
class ErroreImport:
    riga: int
    messaggio: str
    id_ordine: str | None = None
    cliente: str | None = None


@dataclass
class RisultatoImport:
    ordini_creati: int = 0
    errori: list[ErroreImport] = field(default_factory=list)


@dataclass
class AnteprimaImport:
    """Esito di una validazione a secco (RF9, GUI): nessuna riga viene persistita."""

    righe_valide: int = 0
    errori: list[ErroreImport] = field(default_factory=list)


@dataclass
class EsitoValidazioneOrdine:
    ammesso: bool
    motivo: str | None = None


@dataclass
class RisultatoOperazioneViaggio:
    ok: bool
    viaggio_id: str | None = None
    motivo: str | None = None


@dataclass
class RisultatoOperazioneOrdine:
    ok: bool
    ordine_id: str | None = None
    motivo: str | None = None


@dataclass
class OrdineVista:
    id: str
    cliente: str
    indirizzo: str
    negozio_partner: str
    peso: float
    volume_cargo: float
    stato: str
    data_arrivo_viaggio: datetime | None
    # RF16 (GUI): l'icona "Registra esito" e' visibile/abilitata solo per un ordine su un
    # viaggio attualmente IN_CORSO che non ha gia' un EsitoConsegna per QUEL viaggio - stessa
    # condizione verificata (in modo autoritativo) da GestoreRendicontazione.registra_esito(),
    # duplicata qui solo per decidere se mostrare l'azione, non per bypassare la validazione.
    puo_registrare_esito: bool = False


@dataclass
class PaginaOrdini:
    """Pagina di risultati di visualizza_ordini: solo la fetta richiesta + il totale filtrato."""

    ordini: list[OrdineVista]
    totale: int


@dataclass
class ViaggioVista:
    id: str
    squadra_id: str
    n_ordini: int
    data_partenza_prevista: datetime
    data_arrivo_prevista: datetime
    stato: str
    capacita_percentuale: float


@dataclass
class PaginaViaggi:
    """Pagina di risultati di visualizza_viaggi: solo la fetta richiesta + il totale filtrato."""

    viaggi: list[ViaggioVista]
    totale: int


@dataclass
class OrdineDiViaggioVista:
    id: str
    cliente: str
    indirizzo: str
    negozio_partner: str


@dataclass
class DettaglioViaggio:
    """Dettaglio read-only di un Viaggio (GUI: click sull'ID in Viaggi, stesso pattern del
    modale dettaglio di GestoreSquadre.dettaglio_squadra)."""

    id: str
    stato: str
    dipendenti: list[str]
    ordini: list[OrdineDiViaggioVista]


def verifica_idoneita_risorsa(ordine: Ordine, camion: Camion, dipendenti: list[Dipendente]) -> bool:
    """Idoneita' categoria<->risorsa (RF11): sponda idraulica per Big, certificazione gas per
    CertificazioneGas. Include anche RF3/RF6: un camion dismesso o un dipendente licenziato
    (flg_attivo=False, soft delete) non sono mai idonei per un nuovo viaggio qualunque sia la
    categoria dell'ordine - i viaggi gia' pianificati/in corso non vengono toccati retroattivamente."""
    if not camion.flg_attivo or not all(dipendente.flg_attivo for dipendente in dipendenti):
        return False
    if ordine.categoria_consegna == CategoriaConsegna.BIG:
        return camion.flg_sponda_idraulica
    if ordine.categoria_consegna == CategoriaConsegna.CERTIFICAZIONE_GAS:
        return any(dipendente.flg_certificazione_gas for dipendente in dipendenti)
    return True


def valida_ordine_per_viaggio(
    ordine: Ordine,
    camion: Camion,
    dipendenti: list[Dipendente],
    peso_occupato: float,
    volume_occupato: float,
) -> EsitoValidazioneOrdine:
    """Validazione RF11 completa (idoneita' + capacita' residua) con motivo del rifiuto."""
    if not camion.flg_attivo:
        return EsitoValidazioneOrdine(ammesso=False, motivo="Il camion non e' piu' in servizio")
    if not all(dipendente.flg_attivo for dipendente in dipendenti):
        return EsitoValidazioneOrdine(
            ammesso=False, motivo="Un membro della squadra non e' piu' in servizio (licenziato)"
        )
    if not verifica_idoneita_risorsa(ordine, camion, dipendenti):
        if ordine.categoria_consegna == CategoriaConsegna.BIG:
            return EsitoValidazioneOrdine(ammesso=False, motivo=MOTIVO_SPONDA_IDRAULICA_MANCANTE)
        return EsitoValidazioneOrdine(ammesso=False, motivo=MOTIVO_CERTIFICAZIONE_GAS_MANCANTE)
    if peso_occupato + ordine.peso > camion.peso_massimo:
        return EsitoValidazioneOrdine(
            ammesso=False, motivo="Il peso dell'ordine supererebbe la capacita' massima del camion"
        )
    if volume_occupato + ordine.volume_cargo > camion.volume_massimo:
        return EsitoValidazioneOrdine(
            ammesso=False, motivo="Il volume dell'ordine supererebbe la capacita' massima del camion"
        )
    return EsitoValidazioneOrdine(ammesso=True)


def crea_viaggio_persistito(
    session: Session,
    ora_partenza: datetime,
    data_arrivo_prevista: datetime,
    composizione_id: str,
    stato_viaggio: StatoViaggio,
    ordini_ids: tuple[str, ...] = (),
) -> tuple[str, int]:
    """Genera il prossimo id sequenziale V-YYYYMMDD-NN per il giorno e persiste il Viaggio
    (piu' l'aggancio degli ordini indicati, se presenti) sulla session gia' aperta passata come
    parametro. Non apre una sessione propria e non fa commit: il chiamante controlla i confini
    della transazione (necessario per preservare l'atomicita' di un batch multi-viaggio in
    MotoreOttimizzazione.applica_piano).

    Aggancia solo gli ordini ancora disponibili (stato RICEVUTO e viaggio_id None): tra il calcolo
    di un piano e la sua applicazione un ordine candidato puo' essere stato agganciato altrove (es.
    a una bozza manuale RF10), e riassegnarlo qui lo ruberebbe silenziosamente. Restituisce l'id del
    viaggio e il numero di ordini effettivamente agganciati.
    """
    prefisso = f"V-{ora_partenza:%Y%m%d}-"
    progressivo = len(session.scalars(select(Viaggio.id).where(Viaggio.id.like(f"{prefisso}%"))).all()) + 1
    viaggio_id = f"{prefisso}{progressivo:02d}"

    session.add(
        Viaggio(
            id=viaggio_id,
            data_partenza_prevista=ora_partenza,
            data_arrivo_prevista=data_arrivo_prevista,
            data_creazione=datetime.now(),
            km_percorsi=None,
            stato_viaggio=stato_viaggio,
            composizione_id=composizione_id,
        )
    )
    ordini_agganciati = 0
    if ordini_ids:
        ordini_da_agganciare = session.scalars(
            select(Ordine).where(
                Ordine.id.in_(ordini_ids),
                Ordine.stato_ordine == StatoOrdine.RICEVUTO,
                Ordine.viaggio_id.is_(None),
            )
        ).all()
        for o in ordini_da_agganciare:
            o.viaggio_id = viaggio_id
            o.stato_ordine = StatoOrdine.PIANIFICATO
        ordini_agganciati = len(ordini_da_agganciare)
    session.flush()
    return viaggio_id, ordini_agganciati


def _prepara_ordini_da_file(
    percorso_file: Path, negozio_partner: str, id_esistenti: set[str]
) -> tuple[list[Ordine], list[ErroreImport]]:
    """Legge e valida le righe del file (RF9) senza persistere nulla: le righe valide diventano
    `Ordine` pronti per una sessione, le righe scartate finiscono in `ErroreImport` (con
    `id_ordine`/`cliente` quando disponibili, per la tabella di anteprima lato GUI). Riusata sia
    dall'anteprima a secco sia dal commit reale, cosi' la validazione esiste in un solo posto -
    non muta `id_esistenti` del chiamante (lavora su una copia)."""
    if not negozio_partner.strip():
        return [], [ErroreImport(riga=0, messaggio="Negozio partner obbligatorio")]

    if percorso_file.suffix.lower() == ".xlsx":
        header, righe = _righe_da_xlsx(percorso_file)
    else:
        with open(percorso_file, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            header = reader.fieldnames
            righe = list(reader)

    if header != COLONNE_ATTESE:
        return [], [
            ErroreImport(
                riga=0,
                messaggio=f"Header non riconosciuto, attese colonne: {';'.join(COLONNE_ATTESE)}",
            )
        ]

    negozio_partner = negozio_partner.strip()
    id_esistenti = set(id_esistenti)
    ordini_validi: list[Ordine] = []
    errori: list[ErroreImport] = []

    for numero_riga, riga in enumerate(righe, start=2):
        id_ordine = riga["ID_Ordine"]
        cliente = riga.get("Cliente")
        if id_ordine in id_esistenti:
            errori.append(ErroreImport(numero_riga, "ID duplicato", id_ordine, cliente))
            continue

        # TypeError: riga con meno colonne dell'header -> csv.DictReader riempie i campi
        # mancanti con None (float(None)/Enum(None) sollevano TypeError). Un try/except per
        # campo (non uno unico) per dare un motivo di scarto specifico (RF9, tabella di
        # anteprima lato GUI) invece di un messaggio tecnico generico.
        try:
            peso = float(riga["Peso"])
        except (ValueError, TypeError):
            errori.append(ErroreImport(numero_riga, "Peso non valido", id_ordine, cliente))
            continue
        try:
            volume = float(riga["Volume"])
        except (ValueError, TypeError):
            errori.append(ErroreImport(numero_riga, "Volume non valido", id_ordine, cliente))
            continue
        try:
            categoria = CategoriaConsegna(riga["Categoria"])
        except (ValueError, TypeError):
            errori.append(ErroreImport(numero_riga, "Categoria non valida", id_ordine, cliente))
            continue

        parti = [parte.strip() for parte in riga["Indirizzo"].rsplit(",", 1)]
        if len(parti) != 2:
            errori.append(ErroreImport(numero_riga, "Indirizzo non valido", id_ordine, cliente))
            continue
        indirizzo, comune = parti
        if riga["Provincia"] is None:
            # Riga con meno colonne dell'header in cui manca solo l'ultima
            # (Provincia): Peso/Volume/Categoria sono presenti e superano il
            # parsing sopra, quindi il TypeError non scatta e riga["Provincia"]
            # resta None (restval di csv.DictReader). Va scartata, non deve
            # far crashare l'import con AttributeError su None.strip().
            errori.append(ErroreImport(numero_riga, "Provincia mancante", id_ordine, cliente))
            continue
        provincia = riga["Provincia"].strip()
        coordinate = geocodifica_comune(comune, provincia)
        lat, lon = coordinate if coordinate is not None else (None, None)

        ordini_validi.append(
            Ordine(
                id=id_ordine,
                indirizzo=indirizzo,
                comune=comune,
                provincia=provincia,
                lat=lat,
                lon=lon,
                cliente=cliente,
                peso=peso,
                volume_cargo=volume,
                categoria_consegna=categoria,
                stato_ordine=StatoOrdine.RICEVUTO,
                data_importazione=datetime.now(),
                data_consegna=None,
                viaggio_id=None,
                negozio_partner=negozio_partner,
            )
        )
        id_esistenti.add(id_ordine)

    return ordini_validi, errori


class GestoreLogistica:
    def __init__(self, session_factory: sessionmaker = SessionLocal) -> None:
        self.session_factory = session_factory

    def importa_ordini(self, percorso_file: Path, negozio_partner: str) -> RisultatoImport:
        with self.session_factory() as session:
            id_esistenti = set(session.scalars(select(Ordine.id)))
            nuovi_ordini, errori = _prepara_ordini_da_file(percorso_file, negozio_partner, id_esistenti)
            session.add_all(nuovi_ordini)
            session.commit()
            return RisultatoImport(ordini_creati=len(nuovi_ordini), errori=errori)

    def anteprima_import_ordini(self, percorso_file: Path, negozio_partner: str) -> AnteprimaImport:
        """RF9 (anteprima, GUI): valida il file come farebbe `importa_ordini`, ma senza
        scrivere nulla sul DB - usata dallo step "Risultato" del modale Importa CSV prima che
        l'utente confermi con "Importa N ordini"."""
        with self.session_factory() as session:
            id_esistenti = set(session.scalars(select(Ordine.id)))
        nuovi_ordini, errori = _prepara_ordini_da_file(percorso_file, negozio_partner, id_esistenti)
        return AnteprimaImport(righe_valide=len(nuovi_ordini), errori=errori)

    def elenco_negozi_partner(self) -> list[str]:
        """Valori distinti di `negozio_partner` gia' visti (RF9, GUI): suggerimenti per il
        selettore "select o crea nuovo" del modale Importa CSV - non e' un'entita' a se' nel
        modello dati, solo una stringa libera su `Ordine`."""
        with self.session_factory() as session:
            valori = session.scalars(
                select(Ordine.negozio_partner).distinct().where(Ordine.negozio_partner.is_not(None))
            ).all()
            return sorted(v for v in valori if v)

    def importa_ordini_async(self, percorso_file: Path, negozio_partner: str) -> "Future[RisultatoImport]":
        """RNF3: come importa_ordini, ma eseguito su un thread separato per non bloccare la GUI
        durante l'importazione massiva (RF9)."""
        return esegui_in_background(lambda: self.importa_ordini(percorso_file, negozio_partner))

    def avvia_composizione_viaggio(
        self,
        composizione_id: str,
        ora_partenza: datetime,
        durata_prevista: timedelta = timedelta(hours=8),
    ) -> RisultatoOperazioneViaggio:
        """RF10 (avvio): crea un Viaggio in stato IN_COMPOSIZIONE, senza ordini, per una
        ComposizioneSquadra idonea/attiva/libera in quella data. `data_partenza_prevista` e'
        impostata subito (non alla chiusura): e' il campo su cui calcola_piano verifica se una
        composizione e' gia' occupata quel giorno, quindi una bozza senza questo campo sarebbe
        invisibile a quel controllo.
        """
        with self.session_factory() as session:
            composizione = session.get(ComposizioneSquadra, composizione_id)
            if composizione is None:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo=f"Composizione '{composizione_id}' non trovata"
                )
            if not composizione.flg_attiva:
                return RisultatoOperazioneViaggio(ok=False, motivo="Composizione non attiva")

            giorno = ora_partenza.date()
            fuori_validita = composizione.data_inizio_validita.date() > giorno or (
                composizione.data_fine_validita is not None
                and composizione.data_fine_validita.date() < giorno
            )
            if fuori_validita:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo="Composizione non valida in questa data"
                )

            inizio_giorno = datetime(giorno.year, giorno.month, giorno.day)
            fine_giorno = inizio_giorno + timedelta(days=1)
            occupata = session.scalar(
                select(Viaggio.id).where(
                    Viaggio.composizione_id == composizione_id,
                    Viaggio.data_partenza_prevista >= inizio_giorno,
                    Viaggio.data_partenza_prevista < fine_giorno,
                )
            )
            if occupata is not None:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo=MOTIVO_COMPOSIZIONE_OCCUPATA
                )

            viaggio_id, _ = crea_viaggio_persistito(
                session,
                ora_partenza=ora_partenza,
                data_arrivo_prevista=ora_partenza + durata_prevista,
                composizione_id=composizione_id,
                stato_viaggio=StatoViaggio.IN_COMPOSIZIONE,
            )
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def aggiungi_ordine_a_viaggio(self, viaggio_id: str, ordine_id: str) -> EsitoValidazioneOrdine:
        """RF10/RF11 (aggiunta): valida l'ordine candidato contro lo stato attuale del viaggio
        in composizione; se ammesso lo aggancia (viaggio_id + stato_ordine=PIANIFICATO) e
        aggiorna i totali, altrimenti il viaggio resta invariato e viene restituito il motivo.
        """
        with self.session_factory() as session:
            viaggio = session.get(Viaggio, viaggio_id)
            if viaggio is None:
                return EsitoValidazioneOrdine(ammesso=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio.stato_viaggio != StatoViaggio.IN_COMPOSIZIONE:
                return EsitoValidazioneOrdine(
                    ammesso=False, motivo="Il viaggio non e' in fase di composizione"
                )

            ordine_candidato = session.get(Ordine, ordine_id)
            if ordine_candidato is None:
                return EsitoValidazioneOrdine(ammesso=False, motivo=f"Ordine '{ordine_id}' non trovato")
            if ordine_candidato.viaggio_id is not None:
                return EsitoValidazioneOrdine(ammesso=False, motivo="Ordine gia' assegnato a un viaggio")

            composizione = viaggio.composizione
            camion = composizione.camion
            dipendenti = [composizione.dipendente_1, composizione.dipendente_2]
            peso_occupato = sum(o.peso for o in viaggio.ordini)
            volume_occupato = sum(o.volume_cargo for o in viaggio.ordini)

            esito = valida_ordine_per_viaggio(
                ordine_candidato, camion, dipendenti, peso_occupato, volume_occupato
            )
            if not esito.ammesso:
                return esito

            ordine_candidato.viaggio_id = viaggio_id
            ordine_candidato.stato_ordine = StatoOrdine.PIANIFICATO
            session.commit()
            return esito

    def rimuovi_ordine_da_viaggio(self, viaggio_id: str, ordine_id: str) -> RisultatoOperazioneOrdine:
        """Inverso di `aggiungi_ordine_a_viaggio`: sgancia l'ordine dal viaggio in composizione
        e lo riporta a Ricevuto, cosi' torna candidato per un altro viaggio. Stesso vincolo
        IN_COMPOSIZIONE dell'aggiunta - un viaggio Pianificato/InCorso non e' piu' modificabile."""
        with self.session_factory() as session:
            viaggio = session.get(Viaggio, viaggio_id)
            if viaggio is None:
                return RisultatoOperazioneOrdine(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio.stato_viaggio != StatoViaggio.IN_COMPOSIZIONE:
                return RisultatoOperazioneOrdine(
                    ok=False, motivo="Il viaggio non e' in fase di composizione"
                )

            ordine = session.get(Ordine, ordine_id)
            if ordine is None:
                return RisultatoOperazioneOrdine(ok=False, motivo=f"Ordine '{ordine_id}' non trovato")
            if ordine.viaggio_id != viaggio_id:
                return RisultatoOperazioneOrdine(ok=False, motivo="L'ordine non e' agganciato a questo viaggio")

            ordine.viaggio_id = None
            ordine.stato_ordine = StatoOrdine.RICEVUTO
            session.commit()
            return RisultatoOperazioneOrdine(ok=True, ordine_id=ordine_id)

    def chiudi_composizione_viaggio(self, viaggio_id: str) -> RisultatoOperazioneViaggio:
        """RF10 (chiusura): porta il viaggio da IN_COMPOSIZIONE a PIANIFICATO definitivo.
        Richiede almeno un ordine agganciato: un viaggio vuoto non viene salvato come pianificato.
        """
        with self.session_factory() as session:
            viaggio = session.get(Viaggio, viaggio_id)
            if viaggio is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio.stato_viaggio != StatoViaggio.IN_COMPOSIZIONE:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo="Il viaggio non e' in fase di composizione"
                )
            if not viaggio.ordini:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo="Impossibile chiudere un viaggio senza ordini"
                )

            viaggio.stato_viaggio = StatoViaggio.PIANIFICATO
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def verifica_partenze(self, ora_riferimento: datetime | None = None) -> list[str]:
        """RF14: al superamento dell'orario di partenza programmato, porta i viaggi Pianificato
        a InCorso (inibendo cosi' ulteriori modifiche al carico, gia' possibili solo su viaggi
        IN_COMPOSIZIONE) e i relativi Ordini da Pianificato a InConsegna. Pensato per essere
        invocato periodicamente dallo scheduler interno
        (config.ini [scheduler].verifica_partenza_intervallo_minuti). Ritorna gli id dei viaggi avviati.
        """
        ora_riferimento = ora_riferimento or datetime.now()
        with self.session_factory() as session:
            viaggi_da_avviare = session.scalars(
                select(Viaggio).where(
                    Viaggio.stato_viaggio == StatoViaggio.PIANIFICATO,
                    Viaggio.data_partenza_prevista <= ora_riferimento,
                )
            ).all()
            for viaggio in viaggi_da_avviare:
                for ordine in viaggio.ordini:
                    ordine.stato_ordine = StatoOrdine.IN_CONSEGNA
                viaggio.stato_viaggio = StatoViaggio.IN_CORSO
            session.commit()
            return [viaggio.id for viaggio in viaggi_da_avviare]

    def visualizza_ordini(
        self,
        ricerca: str | None = None,
        filtro_stato: str | list[str] = FILTRO_TUTTI,
        filtro_negozio_partner: str | list[str] = FILTRO_TUTTI,
        filtro_data: date | None = None,
        pagina: int = 1,
        dimensione_pagina: int = 20,
        decrescente: bool = False,
    ) -> PaginaOrdini:
        """Elenco filtrato/ordinato/paginato degli ordini. Filtri: ricerca testuale (cliente,
        indirizzo o id), filtro stato (Tutti/Da pianificare/Pianificato/In consegna/Consegnato/
        Fallito), filtro negozio partner (Tutti/uno o piu' valori distinti gia' visti, stesso
        elenco di elenco_negozi_partner()), filtro su un giorno esatto della data di arrivo del
        viaggio agganciato. Ordinamento per data di arrivo del viaggio (non esiste un campo "data
        ordine" nel modello - nessuna RF lo richiede - quindi l'unico riferimento temporale
        disponibile e' quello del viaggio, coerente con la colonna DATA mostrata in tabella).
        Ordini senza viaggio agganciato (data_arrivo_viaggio=None) restano sempre in coda, in
        entrambe le direzioni di ordinamento - un "non ancora pianificato" non ha una posizione
        temporale sensata rispetto agli altri."""
        with self.session_factory() as session:
            righe_grezze = session.execute(
                select(
                    Ordine.id,
                    Ordine.cliente,
                    Ordine.indirizzo,
                    Ordine.comune,
                    Ordine.negozio_partner,
                    Ordine.peso,
                    Ordine.volume_cargo,
                    Ordine.stato_ordine,
                    Viaggio.data_arrivo_prevista,
                    Ordine.viaggio_id,
                    Viaggio.stato_viaggio,
                )
                .outerjoin(Viaggio, Viaggio.id == Ordine.viaggio_id)
                .order_by(Ordine.id)
            ).all()
            # Coppie (ordine_id, viaggio_id) che hanno gia' un EsitoConsegna: un ordine puo'
            # essere stato su piu' viaggi nella sua vita (RF17 lo riaccoda dopo un Fallito), quindi
            # l'idempotenza va verificata per QUESTO viaggio, non per l'ordine in generale - stessa
            # chiave usata da GestoreRendicontazione.registra_esito().
            coppie_con_esito = set(
                session.execute(select(EsitoConsegna.ordine_id, EsitoConsegna.viaggio_id)).all()
            )

            righe = [
                OrdineVista(
                    id=r[0],
                    cliente=r[1],
                    indirizzo=f"{r[2]}, {r[3]}",
                    negozio_partner=r[4] or "Non specificato",
                    peso=r[5],
                    volume_cargo=r[6],
                    stato=STATO_ORDINE_LABELS[r[7]],
                    data_arrivo_viaggio=r[8],
                    puo_registrare_esito=(
                        r[9] is not None
                        and r[10] == StatoViaggio.IN_CORSO
                        and (r[0], r[9]) not in coppie_con_esito
                    ),
                )
                for r in righe_grezze
            ]

            con_data = sorted(
                (r for r in righe if r.data_arrivo_viaggio is not None),
                key=lambda r: r.data_arrivo_viaggio,
                reverse=decrescente,
            )
            senza_data = [r for r in righe if r.data_arrivo_viaggio is None]
            righe = con_data + senza_data

            valori_stato = _normalizza_filtro_multiplo(filtro_stato, FILTRO_TUTTI)
            if valori_stato:
                righe = [r for r in righe if r.stato in valori_stato]

            valori_negozio_partner = _normalizza_filtro_multiplo(filtro_negozio_partner, FILTRO_TUTTI)
            if valori_negozio_partner:
                righe = [r for r in righe if r.negozio_partner in valori_negozio_partner]

            if filtro_data is not None:
                righe = [
                    r
                    for r in righe
                    if r.data_arrivo_viaggio is not None and r.data_arrivo_viaggio.date() == filtro_data
                ]

            if ricerca:
                termine = ricerca.strip().lower()
                if termine:
                    righe = [
                        r
                        for r in righe
                        if termine in r.id.lower()
                        or termine in r.cliente.lower()
                        or termine in r.indirizzo.lower()
                    ]

            totale = len(righe)
            if dimensione_pagina > 0:
                inizio = max(pagina - 1, 0) * dimensione_pagina
                righe = righe[inizio : inizio + dimensione_pagina]

            return PaginaOrdini(ordini=righe, totale=totale)

    def visualizza_viaggi(
        self,
        ricerca: str | None = None,
        filtro_stato: str | list[str] = FILTRO_TUTTI,
        filtro_data: date | None = None,
        pagina: int = 1,
        dimensione_pagina: int = 20,
        decrescente: bool = False,
        ordina_per: str = ORDINA_PER_PARTENZA,
    ) -> PaginaViaggi:
        """Elenco filtrato/ordinato/paginato dei viaggi. Filtri: ricerca testuale (id viaggio o
        squadra), filtro stato (Tutti/In composizione/Pianificato/In corso/Completato/Annullato),
        filtro su un giorno esatto di data_partenza_prevista, ordinamento su partenza O arrivo
        (entrambe le colonne sono sortable nel mockup, non solo una come per Dipendenti/Camion),
        paginazione server-side. N. ordini e capacita' occupata calcolati con un'unica query
        aggregata (func.count/func.sum via outerjoin, stessa tecnica di dettaglio_squadra in
        gestore_squadre.py - niente N+1). capacita_percentuale e' il maggiore tra peso% e volume%
        occupati sul camion della composizione (il collo di bottiglia del viaggio) - stessa
        formula e stessa colonna "Capacità" gia' usata dalla Proposed Trips Table di
        Pianificazione (vedi gui/pianificazione/pianificazione_data.py:costruisci_righe_piano),
        qui riusata cosi' com'e' invece di reinventarla."""
        with self.session_factory() as session:
            composizione_info: dict[str, tuple[str, float, float]] = {
                r[0]: (r[1], r[2], r[3])
                for r in session.execute(
                    select(
                        ComposizioneSquadra.id_composizione,
                        ComposizioneSquadra.squadra_id,
                        Camion.peso_massimo,
                        Camion.volume_massimo,
                    ).join(Camion, Camion.id == ComposizioneSquadra.camion_id)
                ).all()
            }

            righe_grezze = session.execute(
                select(
                    Viaggio.id,
                    Viaggio.composizione_id,
                    Viaggio.stato_viaggio,
                    Viaggio.data_partenza_prevista,
                    Viaggio.data_arrivo_prevista,
                    func.count(Ordine.id),
                    func.coalesce(func.sum(Ordine.peso), 0.0),
                    func.coalesce(func.sum(Ordine.volume_cargo), 0.0),
                )
                .outerjoin(Ordine, Ordine.viaggio_id == Viaggio.id)
                .group_by(
                    Viaggio.id,
                    Viaggio.composizione_id,
                    Viaggio.stato_viaggio,
                    Viaggio.data_partenza_prevista,
                    Viaggio.data_arrivo_prevista,
                )
            ).all()

            righe: list[ViaggioVista] = []
            for r in righe_grezze:
                squadra_id, peso_massimo, volume_massimo = composizione_info.get(r[1], ("—", 0.0, 0.0))
                peso_pct = (r[6] / peso_massimo * 100) if peso_massimo else 0.0
                volume_pct = (r[7] / volume_massimo * 100) if volume_massimo else 0.0
                righe.append(
                    ViaggioVista(
                        id=r[0],
                        squadra_id=squadra_id,
                        n_ordini=r[5],
                        data_partenza_prevista=r[3],
                        data_arrivo_prevista=r[4],
                        stato=STATO_VIAGGIO_LABELS[r[2]],
                        capacita_percentuale=max(peso_pct, volume_pct),
                    )
                )

            campo_ordinamento = (
                (lambda r: r.data_arrivo_prevista)
                if ordina_per == ORDINA_PER_ARRIVO
                else (lambda r: r.data_partenza_prevista)
            )
            righe.sort(key=campo_ordinamento, reverse=decrescente)

            valori_stato = _normalizza_filtro_multiplo(filtro_stato, FILTRO_TUTTI)
            if valori_stato:
                righe = [r for r in righe if r.stato in valori_stato]

            if filtro_data is not None:
                righe = [r for r in righe if r.data_partenza_prevista.date() == filtro_data]

            if ricerca:
                termine = ricerca.strip().lower()
                if termine:
                    righe = [
                        r
                        for r in righe
                        if termine in r.id.lower() or termine in r.squadra_id.lower()
                    ]

            totale = len(righe)
            if dimensione_pagina > 0:
                inizio = max(pagina - 1, 0) * dimensione_pagina
                righe = righe[inizio : inizio + dimensione_pagina]

            return PaginaViaggi(viaggi=righe, totale=totale)

    def dettaglio_viaggio(self, viaggio_id: str) -> DettaglioViaggio | None:
        """Dettaglio read-only di un Viaggio (GUI: click sull'ID in Viaggi): i due dipendenti
        della composizione agganciata e gli ordini caricati su questo viaggio (id/cliente/
        indirizzo). Ritorna None se il viaggio non esiste. Stesso pattern a due alias di
        GestoreSquadre._select_composizioni_attive per i due FK dipendente_1/dipendente_2 sulla
        stessa tabella."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return None

            dipendente_1 = aliased(Dipendente)
            dipendente_2 = aliased(Dipendente)
            riga_dipendenti = session.execute(
                select(dipendente_1.nome, dipendente_1.cognome, dipendente_2.nome, dipendente_2.cognome)
                .select_from(ComposizioneSquadra)
                .join(dipendente_1, dipendente_1.id == ComposizioneSquadra.dipendente_1_id)
                .join(dipendente_2, dipendente_2.id == ComposizioneSquadra.dipendente_2_id)
                .where(ComposizioneSquadra.id_composizione == viaggio_obj.composizione_id)
            ).first()
            dipendenti = (
                [f"{riga_dipendenti[0]} {riga_dipendenti[1]}", f"{riga_dipendenti[2]} {riga_dipendenti[3]}"]
                if riga_dipendenti is not None
                else []
            )

            ordini_righe = session.execute(
                select(
                    Ordine.id, Ordine.cliente, Ordine.indirizzo, Ordine.comune, Ordine.negozio_partner
                )
                .where(Ordine.viaggio_id == viaggio_id)
                .order_by(Ordine.id)
            ).all()
            ordini = [
                OrdineDiViaggioVista(
                    id=r[0],
                    cliente=r[1],
                    indirizzo=f"{r[2]}, {r[3]}",
                    negozio_partner=r[4] or "Non specificato",
                )
                for r in ordini_righe
            ]

            return DettaglioViaggio(
                id=viaggio_obj.id,
                stato=STATO_VIAGGIO_LABELS[viaggio_obj.stato_viaggio],
                dipendenti=dipendenti,
                ordini=ordini,
            )

    def annulla_viaggio(self, viaggio_id: str) -> RisultatoOperazioneViaggio:
        """Annulla un viaggio (soft-mark ANNULLATO), consentito anche da IN_CORSO - non solo da
        IN_COMPOSIZIONE/PIANIFICATO - bloccato solo se gia' COMPLETATO o gia' ANNULLATO (decisione
        esplicita dell'utente, diverso dal blocco piu' stretto di licenzia_dipendente/
        disattiva_camion). Gli Ordine agganciati tornano RICEVUTO con viaggio_id=None, stesso
        comportamento gia' usato in GestoreRendicontazione.registra_esito per un esito Fallito
        (pattern esistente riusato, non reinventato)."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio_obj.stato_viaggio in (StatoViaggio.COMPLETATO, StatoViaggio.ANNULLATO):
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo=f"Impossibile annullare: viaggio gia' {STATO_VIAGGIO_LABELS[viaggio_obj.stato_viaggio].lower()}",
                )

            for o in viaggio_obj.ordini:
                o.stato_ordine = StatoOrdine.RICEVUTO
                o.viaggio_id = None

            viaggio_obj.stato_viaggio = StatoViaggio.ANNULLATO
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def modifica_date_viaggio(
        self,
        viaggio_id: str,
        data_partenza_prevista: datetime | None = None,
        data_arrivo_prevista: datetime | None = None,
    ) -> RisultatoOperazioneViaggio:
        """Modifica non presente nel mockup Sketch: non esiste un artboard "Viaggi — Modifica
        (modale)" ne' un'operazione di modifica viaggio nelle RF - l'icona matita in tabella e'
        disegnata ma senza specifica. Su decisione esplicita dell'utente, l'icona apre un modale
        che permette di correggere le due date previste, la squadra (modifica_squadra_viaggio) e
        gli ordini caricati (aggiungi_ordine_a_viaggio). Bloccata se il viaggio e' gia' COMPLETATO
        o ANNULLATO (stesso principio di annulla_viaggio: uno stato terminale non si corregge
        piu')."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio_obj.stato_viaggio in (StatoViaggio.COMPLETATO, StatoViaggio.ANNULLATO):
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo=f"Impossibile modificare: viaggio gia' {STATO_VIAGGIO_LABELS[viaggio_obj.stato_viaggio].lower()}",
                )

            if data_partenza_prevista is not None:
                viaggio_obj.data_partenza_prevista = data_partenza_prevista
            if data_arrivo_prevista is not None:
                viaggio_obj.data_arrivo_prevista = data_arrivo_prevista

            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def modifica_squadra_viaggio(self, viaggio_id: str, composizione_id: str) -> RisultatoOperazioneViaggio:
        """Riassegna la composizione (squadra+camion+2 dipendenti) di un Viaggio gia' esistente -
        controparte di modifica_date_viaggio per lo stesso modale "Modifica". Non esisteva prima:
        Viaggio.composizione_id era scrivibile solo alla creazione (crea_viaggio_persistito).

        Consentita solo IN_COMPOSIZIONE/PIANIFICATO (decisione esplicita dell'utente, stesso
        vincolo con cui la GUI nasconde l'intero pulsante "Modifica" per le altre righe - qui
        duplicato lato backend come difesa in profondita', stesso principio delle altre operazioni
        di questo modulo). Se il viaggio ha gia' ordini caricati, la nuova composizione deve essere
        idonea e capiente per TUTTI loro (stessa validazione di valida_ordine_per_viaggio, qui
        applicata in blocco anziche' incrementalmente su un singolo ordine candidato) - altrimenti
        il cambio viene rifiutato per non lasciare ordini gia' caricati su un camion/squadra non
        piu' idonei per loro."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio_obj.stato_viaggio not in (StatoViaggio.IN_COMPOSIZIONE, StatoViaggio.PIANIFICATO):
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo=f"Impossibile modificare: viaggio gia' {STATO_VIAGGIO_LABELS[viaggio_obj.stato_viaggio].lower()}",
                )

            nuova_composizione = session.get(ComposizioneSquadra, composizione_id)
            if nuova_composizione is None:
                return RisultatoOperazioneViaggio(
                    ok=False, motivo=f"Composizione '{composizione_id}' non trovata"
                )
            if not nuova_composizione.flg_attiva:
                return RisultatoOperazioneViaggio(ok=False, motivo="La squadra selezionata non e' attiva")

            camion = nuova_composizione.camion
            dipendenti = [nuova_composizione.dipendente_1, nuova_composizione.dipendente_2]
            if not camion.flg_attivo or not all(dipendente.flg_attivo for dipendente in dipendenti):
                return RisultatoOperazioneViaggio(
                    ok=False, motivo="Il camion o un dipendente della squadra selezionata non e' piu' in servizio"
                )

            peso_totale = sum(o.peso for o in viaggio_obj.ordini)
            volume_totale = sum(o.volume_cargo for o in viaggio_obj.ordini)
            if peso_totale > camion.peso_massimo or volume_totale > camion.volume_massimo:
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo="Il camion della squadra selezionata non ha capacita' sufficiente per gli ordini gia' caricati",
                )
            for ordine_caricato in viaggio_obj.ordini:
                if not verifica_idoneita_risorsa(ordine_caricato, camion, dipendenti):
                    return RisultatoOperazioneViaggio(
                        ok=False,
                        motivo=f"La squadra selezionata non e' idonea per l'ordine '{ordine_caricato.id}' gia' caricato",
                    )

            viaggio_obj.composizione_id = composizione_id
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def ripristina_viaggio(self, viaggio_id: str) -> RisultatoOperazioneViaggio:
        """Annulla un annullamento fatto per errore. Non esiste uno stato "precedente" salvato
        (poteva essere IN_COMPOSIZIONE, PIANIFICATO o IN_CORSO) e gli Ordine staccati da
        annulla_viaggio() non vengono ricollegati automaticamente (potrebbero nel frattempo essere
        stati presi da un altro viaggio) - su decisione esplicita dell'utente il viaggio torna
        sempre a IN_COMPOSIZIONE, l'unico stato che ammette 0 ordini senza violare l'invariante di
        chiudi_composizione_viaggio ("richiede almeno un ordine" per passare a PIANIFICATO).
        L'operatore poi ri-aggancia gli ordini a mano con aggiungi_ordine_a_viaggio()."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")
            if viaggio_obj.stato_viaggio != StatoViaggio.ANNULLATO:
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo=f"Impossibile ripristinare: viaggio non annullato (gia' {STATO_VIAGGIO_LABELS[viaggio_obj.stato_viaggio].lower()})",
                )

            viaggio_obj.stato_viaggio = StatoViaggio.IN_COMPOSIZIONE
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def elimina_viaggio_definitivamente(self, viaggio_id: str) -> RisultatoOperazioneViaggio:
        """Hard-delete, irreversibile: rimuove il viaggio dal database. Rifiuta se ha ordini
        agganciati (anche storici: un Ordine consegnato mantiene per sempre il suo viaggio_id come
        record) o esiti di consegna registrati - altrimenti romperebbe il vincolo di integrita'
        referenziale con quelle righe."""
        with self.session_factory() as session:
            viaggio_obj = session.get(Viaggio, viaggio_id)
            if viaggio_obj is None:
                return RisultatoOperazioneViaggio(ok=False, motivo=f"Viaggio '{viaggio_id}' non trovato")

            ordine_bloccante = session.scalar(select(Ordine.id).where(Ordine.viaggio_id == viaggio_id))
            if ordine_bloccante is not None:
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo="Impossibile eliminare definitivamente: il viaggio ha ordini agganciati",
                )

            esito_bloccante = session.scalar(
                select(EsitoConsegna.id).where(EsitoConsegna.viaggio_id == viaggio_id)
            )
            if esito_bloccante is not None:
                return RisultatoOperazioneViaggio(
                    ok=False,
                    motivo="Impossibile eliminare definitivamente: il viaggio ha esiti di consegna registrati",
                )

            session.delete(viaggio_obj)
            session.commit()
            return RisultatoOperazioneViaggio(ok=True, viaggio_id=viaggio_id)

    def elimina_ordine_definitivamente(self, ordine_id: str) -> RisultatoOperazioneOrdine:
        """Hard-delete, irreversibile: rimuove l'ordine dal database. Rifiuta se ha esiti di
        consegna registrati o e' incluso in un report consuntivo - altrimenti romperebbe il vincolo
        di integrita' referenziale con quelle righe."""
        with self.session_factory() as session:
            ordine_obj = session.get(Ordine, ordine_id)
            if ordine_obj is None:
                return RisultatoOperazioneOrdine(ok=False, motivo=f"Ordine '{ordine_id}' non trovato")

            esito_bloccante = session.scalar(
                select(EsitoConsegna.id).where(EsitoConsegna.ordine_id == ordine_id)
            )
            if esito_bloccante is not None:
                return RisultatoOperazioneOrdine(
                    ok=False,
                    motivo="Impossibile eliminare definitivamente: l'ordine ha esiti di consegna registrati",
                )

            report_bloccante = session.scalar(
                select(report_ordini.c.report_id).where(report_ordini.c.ordine_id == ordine_id)
            )
            if report_bloccante is not None:
                return RisultatoOperazioneOrdine(
                    ok=False,
                    motivo="Impossibile eliminare definitivamente: l'ordine e' incluso in un report consuntivo",
                )

            session.delete(ordine_obj)
            session.commit()
            return RisultatoOperazioneOrdine(ok=True, ordine_id=ordine_id)
